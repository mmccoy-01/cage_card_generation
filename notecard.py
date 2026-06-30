from __future__ import annotations

import argparse
import io
import math
import re
import textwrap
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any, BinaryIO

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

# ReportLab is only needed when a PDF is generated. Import it defensively so
# app startup does not crash on hosted systems where dependencies have not
# installed yet; the app will show a clear build error instead.
REPORTLAB_IMPORT_ERROR: Exception | None = None
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas
except Exception as exc:  # pragma: no cover - only used when dependency is missing
    REPORTLAB_IMPORT_ERROR = exc

    class _FallbackColors:
        black = "#000000"
        white = "#FFFFFF"

        @staticmethod
        def HexColor(value: str) -> str:
            return value

    class _FallbackCanvasModule:
        class Canvas:  # type: ignore[no-redef]
            pass

    colors = _FallbackColors()  # type: ignore[assignment]
    canvas = _FallbackCanvasModule()  # type: ignore[assignment]
    mm = 72.0 / 25.4
    letter = (612.0, 792.0)

    def landscape(page_size: tuple[float, float]) -> tuple[float, float]:
        return (page_size[1], page_size[0])


MAX_MICE_PER_CAGE = 6
PRINT_COLUMN = "Print Card?"

# Physical cage-card target requested by the user.
CARD_WIDTH_MM = 127.0
CARD_HEIGHT_MM = 77.0
CARD_W = CARD_WIDTH_MM * mm
CARD_H = CARD_HEIGHT_MM * mm

PAGE_SIZE = landscape(letter)
PAGE_W, PAGE_H = PAGE_SIZE
CARDS_PER_PAGE = 4
GUTTER_W = 6 * mm
GUTTER_H = 6 * mm
CARD_LEFT_MARGIN = (PAGE_W - (2 * CARD_W + GUTTER_W)) / 2
CARD_TOP_MARGIN = (PAGE_H - (2 * CARD_H + GUTTER_H)) / 2
LEFT_CARD_X = CARD_LEFT_MARGIN
RIGHT_CARD_X = CARD_LEFT_MARGIN + CARD_W + GUTTER_W
TOP_CARD_Y = PAGE_H - CARD_TOP_MARGIN
BOTTOM_CARD_Y = TOP_CARD_Y - CARD_H - GUTTER_H
CARD_SLOTS = [
    (LEFT_CARD_X, TOP_CARD_Y),
    (RIGHT_CARD_X, TOP_CARD_Y),
    (LEFT_CARD_X, BOTTOM_CARD_Y),
    (RIGHT_CARD_X, BOTTOM_CARD_Y),
]

# Back-page slot mapping for landscape duplex printing with "flip on long edge".
# The user's printer places backs on the opposite card horizontally within
# each row, so we compensate by mirroring left/right on the back page.
#
# Front slots:
#   0 = top-left
#   1 = top-right
#   2 = bottom-left
#   3 = bottom-right
#
# Back page placement:
#   front 0 cage -> back slot 1
#   front 1 cage -> back slot 0
#   front 2 cage -> back slot 3
#   front 3 cage -> back slot 2
BACK_SLOT_FOR_FRONT_SLOT = {
    0: 1,
    1: 0,
    2: 3,
    3: 2,
}

HEADER_NAMES = {
    "print_card": ["print card?", "print card", "print", "include", "selected"],
    "cage_tag": ["cage tag"],
    "num_mice": ["# of mice", "num mice", "number of mice"],
    "disposition": ["disposition"],  # optional backward compatibility only
    "mating_sid": ["mating sid", "mating id", "mating", "sid"],
    "litter_mouseline": [
        "litter mouseline",
        "litter mouse line",
        "litter strain",
        "litter line",
    ],
    "mice_tags": ["mice tags [sex, dob, age]", "mice tags", "mouse tags"],
    "genotypes": ["genotypes", "genotype"],
    "comment": ["comment", "comments", "notes"],
    "source": ["source"],
    "protocol_num": ["protocol", "protocol number", "protocol #", "protocol num"],
    "approved_date": ["approved", "approved date", "approval date"],
    "expires_date": ["expires", "expires date", "expiration date", "expiry date"],
}

MATING_HEADER_NAMES = {
    "cage_tag": ["cage tag", "mating cage"],
    "comment": ["comment", "comments", "notes"],
    "mating_sid": ["mating sid", "mating id", "mating", "sid"],
    "litter_mouseline": ["litter mouseline", "litter mouse line", "litter line"],
    "m_genotype": ["m genotype", "male genotype", "sire genotype"],
    "f_genotype": ["f genotype", "female genotype", "dam genotype"],
    "m_tag": ["m tag", "male tag", "sire tag"],
    "f_tag": ["f tag", "female tag", "dam tag"],
    "num_litters": ["# litters", "num litters", "number litters", "number of litters"],
    "litter_history": [
        "litter sids [size, dob, wean date, state, end date, end reason]",
        "litter sids",
        "litter history",
    ],
}

DEFAULT_SETTINGS = {
    "PI_name": "",
    "protocol_num": "",
    "approved_date": "",
    "expires_date": "",
    "contact_name": "",
    "contact_phone": "",
    "species": "Mouse",
    "source": "",
}


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_display(value: Any, separator: str = " / ") -> str:
    text = safe_str(value)
    if not text:
        return ""
    lines = [line.strip() for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    lines = [line for line in lines if line]
    return separator.join(lines)


def normalize_sid(value: Any) -> str:
    text = safe_str(value)
    if not text:
        return ""
    match = re.match(r"^([0-9]+)(?:\.0+)?$", text)
    if match:
        return match.group(1)
    return text


def normalize_sid_list(value: Any) -> list[str]:
    """
    Return one or more normalized Mating SIDs from a SoftMouse cell.

    SoftMouse cells can contain historical/current Mating SIDs on separate
    lines, for example "166\n178". Treat those as separate possible IDs
    rather than one combined ID. Numeric Excel values such as 178.0 are
    normalized to "178".
    """
    text = safe_str(value)
    if not text:
        return []

    # Prefer explicit numeric IDs because Mating SID is expected to be numeric.
    # This handles newline-, comma-, slash-, and semicolon-separated cells.
    numeric_ids = re.findall(r"\d+(?:\.0+)?", text)
    if numeric_ids:
        out: list[str] = []
        for item in numeric_ids:
            sid = normalize_sid(item)
            if sid and sid not in out:
                out.append(sid)
        return out

    # Fallback for non-numeric IDs, preserving order and dropping blanks.
    parts = re.split(r"[\r\n,;/]+", text)
    out = []
    for part in parts:
        sid = normalize_sid(part.strip())
        if sid and sid not in out:
            out.append(sid)
    return out


def choose_mating_sid(mating_sids: list[str], mating_lookup: dict[str, dict[str, str]]) -> str:
    """Choose the best SID to use for sire/dam lookup.

    If there are multiple SIDs in the cage workbook, the most recent/current
    one is usually the last value listed. Choose the last SID that exists in
    the mating workbook; otherwise fall back to the last listed SID.
    """
    if not mating_sids:
        return ""
    for sid in reversed(mating_sids):
        if sid in mating_lookup:
            return sid
    return mating_sids[-1]


def normalize_settings(settings: dict[str, Any] | None) -> dict[str, str]:
    normalized = dict(DEFAULT_SETTINGS)
    if settings:
        normalized.update({k: safe_str(v) for k, v in settings.items()})
    if not normalized["species"]:
        normalized["species"] = "Mouse"
    return normalized


def normalize_header(value: Any) -> str:
    text = safe_str(value).lower()
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"^\$+\s*", "", text)
    text = re.sub(r"[_\.]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def build_header_index(header_row: list[Any], names: dict[str, list[str]]) -> dict[str, int | None]:
    normalized = {normalize_header(v): i for i, v in enumerate(header_row) if safe_str(v)}
    out: dict[str, int | None] = {}
    for key, candidates in names.items():
        out[key] = None
        for name in candidates:
            if normalize_header(name) in normalized:
                out[key] = normalized[normalize_header(name)]
                break
    return out


def cell(data_row: list[Any], header_index: dict[str, int | None], key: str, default: Any = "") -> Any:
    idx = header_index.get(key)
    if idx is None or idx >= len(data_row):
        return default
    value = data_row[idx]
    return default if value is None else value


def cleaned_lines(value: Any, keep_blank_lines: bool = False) -> list[str]:
    text = safe_str(value)
    if not text:
        return [""] if keep_blank_lines else []
    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    if keep_blank_lines:
        return [line.strip() for line in lines]
    return [line.strip() for line in lines if line.strip()]


def safe_int(value: Any, default: int = 0) -> int:
    text = safe_str(value)
    if not text:
        return default
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else default


def safe_bool(value: Any, default: bool = True) -> bool:
    text = safe_str(value).lower()
    if not text:
        return default
    if text in {"true", "t", "yes", "y", "1", "include", "print", "selected", "x"}:
        return True
    if text in {"false", "f", "no", "n", "0", "skip", "exclude", "unchecked"}:
        return False
    return default


def format_date(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(value).strftime("%Y-%m-%d")
        except Exception:
            pass
    text = safe_str(value)
    if not text:
        return ""
    # Use the first non-empty line and ignore SoftMouse's angle-bracket older values.
    first = next((line.strip() for line in text.splitlines() if line.strip()), "")
    first = re.sub(r"<.*?>", "", first).strip()
    if not first:
        return ""
    parsed = pd.to_datetime(first, errors="coerce")
    if pd.notna(parsed):
        return parsed.strftime("%Y-%m-%d")
    return first


def parse_mouse_lines(mouse_lines: list[str]) -> list[dict[str, str]]:
    parsed: list[dict[str, str]] = []
    for raw in mouse_lines:
        tag = raw.split("[")[0].strip()
        sex_match = re.search(r"\[(M|F)", raw, flags=re.IGNORECASE)
        dob_match = re.search(r"([0-1]?[0-9][-/][0-3]?[0-9][-/](?:20)?[0-9]{2})", raw)
        parsed.append(
            {
                "tag": tag,
                "sex": sex_match.group(1).upper() if sex_match else "",
                "dob": format_date(dob_match.group(1)) if dob_match else "",
                "raw": raw,
            }
        )
    return parsed


def compact_note(comment: str, overflow_count: int = 0) -> str:
    parts: list[str] = []
    note = " ".join(cleaned_lines(comment))
    if note:
        parts.append(note)
    if overflow_count > 0:
        parts.append(f"+{overflow_count} more mouse(s) not shown")
    if not parts:
        return ""
    return textwrap.shorten(" | ".join(parts), width=95, placeholder="...")


def _rows_from_source(source: str | Path | bytes | BinaryIO | pd.DataFrame) -> tuple[list[list[Any]], list[str]]:
    captured_warnings: list[str] = []
    if isinstance(source, pd.DataFrame):
        df = source.fillna("")
        return [list(df.columns)] + df.values.tolist(), captured_warnings

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        wb = load_workbook(source, data_only=True)
    for warning_obj in caught:
        message = str(warning_obj.message)
        if "Workbook contains no default style" not in message:
            captured_warnings.append(message)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)], captured_warnings


def load_mating_lookup(
    mating_source: str | Path | bytes | BinaryIO | pd.DataFrame | None,
) -> tuple[dict[str, dict[str, str]], list[str]]:
    warnings_out: list[str] = []
    if mating_source is None:
        return {}, warnings_out

    rows, captured = _rows_from_source(mating_source)
    warnings_out.extend(captured)
    if not rows:
        return {}, warnings_out

    header_index = build_header_index(rows[0], MATING_HEADER_NAMES)
    if header_index.get("mating_sid") is None:
        warnings_out.append("Mating workbook was uploaded, but no Mating SID column was found.")
        return {}, warnings_out

    lookup: dict[str, dict[str, str]] = {}
    for raw in rows[1:]:
        sid = normalize_sid(cell(raw, header_index, "mating_sid"))
        if not sid:
            continue
        lookup[sid] = {
            "mating_sid": sid,
            "mating_cage": clean_display(cell(raw, header_index, "cage_tag")),
            "comment": clean_display(cell(raw, header_index, "comment"), separator="\n"),
            "litter_mouseline": clean_display(cell(raw, header_index, "litter_mouseline")),
            "sire_tag": clean_display(cell(raw, header_index, "m_tag")),
            "dam_tag": clean_display(cell(raw, header_index, "f_tag")),
            "sire_genotype": clean_display(cell(raw, header_index, "m_genotype"), separator="\n"),
            "dam_genotype": clean_display(cell(raw, header_index, "f_genotype"), separator="\n"),
            "num_litters": safe_str(cell(raw, header_index, "num_litters")),
            "litter_history": clean_display(cell(raw, header_index, "litter_history"), separator="\n"),
        }
    return lookup, warnings_out


def load_cages(
    xlsx_source: str | Path | bytes | BinaryIO | pd.DataFrame,
    mating_source: str | Path | bytes | BinaryIO | pd.DataFrame | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    rows, captured_warnings = _rows_from_source(xlsx_source)

    if not rows:
        return [], captured_warnings

    header_index = build_header_index(rows[0], HEADER_NAMES)
    required_headers = ["cage_tag", "num_mice", "mice_tags", "genotypes"]
    missing = [name for name in required_headers if header_index.get(name) is None]
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    mating_lookup, mating_warnings = load_mating_lookup(mating_source)
    captured_warnings.extend(mating_warnings)

    data_rows = rows[1:]
    cages: list[dict[str, Any]] = []

    for raw in data_rows:
        if header_index.get("print_card") is not None and not safe_bool(cell(raw, header_index, "print_card", True), default=True):
            continue

        cage_tag = safe_str(cell(raw, header_index, "cage_tag"))
        mating_sids = normalize_sid_list(cell(raw, header_index, "mating_sid"))
        mating_sid = choose_mating_sid(mating_sids, mating_lookup)
        litter_mouseline = clean_display(cell(raw, header_index, "litter_mouseline"))
        disposition_raw = safe_str(cell(raw, header_index, "disposition")).lower()
        disposition = "mating" if mating_sids or litter_mouseline or disposition_raw == "mating" else "stock"

        if not cage_tag and not mating_sids and not litter_mouseline:
            continue

        declared_num = safe_int(cell(raw, header_index, "num_mice", 0))
        mouse_lines = cleaned_lines(cell(raw, header_index, "mice_tags"))
        mice = parse_mouse_lines(mouse_lines)
        genotype_lines = cleaned_lines(cell(raw, header_index, "genotypes"), keep_blank_lines=True)

        # Do not warn when there are fewer mouse-tag lines than the declared
        # cage count. That is expected for ungenotyped/untagged pups: # of mice
        # represents the physical cage count, while mouse-tag lines represent
        # the mice that can be printed individually on the card. Still warn if
        # there are more tag lines than the declared cage count, because that
        # is more likely to indicate a data-entry/export issue.
        if declared_num and len(mice) > declared_num:
            captured_warnings.append(
                f"Cage {cage_tag or '(blank)'} says {declared_num} mice, but {len(mice)} mouse-tag line(s) were found."
            )

        mating_info = mating_lookup.get(mating_sid, {}) if mating_sid else {}
        if disposition == "mating" and mating_source is not None and mating_sids and not mating_info:
            sid_text = ", ".join(mating_sids)
            captured_warnings.append(
                f"Cage {cage_tag or '(blank)'} has Mating SID(s) {sid_text}, but none matched a row in the mating workbook."
            )

        if disposition == "mating" and not litter_mouseline:
            litter_mouseline = safe_str(mating_info.get("litter_mouseline", ""))

        cages.append(
            {
                "cage_tag": cage_tag,
                "source": safe_str(cell(raw, header_index, "source")),
                "protocol_num": safe_str(cell(raw, header_index, "protocol_num")),
                "approved_date": format_date(cell(raw, header_index, "approved_date")),
                "expires_date": format_date(cell(raw, header_index, "expires_date")),
                "disposition": disposition,
                "mating_sid": mating_sid,
                "mating_sids": mating_sids,
                "litter_mouseline": litter_mouseline,
                "mice": mice,
                "genotypes": genotype_lines,
                "comment": safe_str(cell(raw, header_index, "comment")),
                "mating_info": mating_info,
            }
        )

    return cages, captured_warnings


# ----------------------------- PDF drawing -----------------------------

def _hex(value: str) -> colors.Color:
    return colors.HexColor(value)


HEADER_FILL = _hex("#D9EAD3")
LABEL_FILL = _hex("#F2F2F2")
TABLE_FILL = _hex("#EDEDED")
BLACK = colors.black
WHITE = colors.white


def _wrap_text(c: canvas.Canvas, text: str, font_name: str, font_size: float, max_width: float) -> list[str]:
    text = safe_str(text)
    if not text:
        return []
    lines: list[str] = []
    for raw_line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        words = raw_line.split()
        if not words:
            lines.append("")
            continue
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            if c.stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        # Break very long unspaced chunks if needed.
        if c.stringWidth(current, font_name, font_size) <= max_width:
            lines.append(current)
        else:
            chunk = ""
            for char in current:
                if c.stringWidth(chunk + char, font_name, font_size) <= max_width:
                    chunk += char
                else:
                    if chunk:
                        lines.append(chunk)
                    chunk = char
            if chunk:
                lines.append(chunk)
    return lines


def draw_cell(
    c: canvas.Canvas,
    x: float,
    y_top: float,
    w: float,
    h: float,
    text: Any = "",
    *,
    font: str = "Helvetica",
    size: float = 7.0,
    bold: bool = False,
    align: str = "left",
    valign: str = "middle",
    fill: colors.Color | None = None,
    text_color: colors.Color = BLACK,
    border: bool = True,
    wrap: bool = True,
    pad: float = 1.3 * mm,
) -> None:
    y = y_top - h
    if fill is not None:
        c.setFillColor(fill)
        c.rect(x, y, w, h, fill=1, stroke=0)
    if border:
        c.setStrokeColor(BLACK)
        c.setLineWidth(0.45)
        c.rect(x, y, w, h, fill=0, stroke=1)

    text_str = safe_str(text)
    if not text_str:
        return

    font_name = "Helvetica-Bold" if bold else font
    c.setFont(font_name, size)
    c.setFillColor(text_color)

    max_width = max(1, w - 2 * pad)
    if wrap:
        lines = _wrap_text(c, text_str, font_name, size, max_width)
    else:
        lines = [text_str]

    line_height = size * 1.15
    max_lines = max(1, int((h - 1.5 * mm) // line_height))
    if len(lines) > max_lines:
        lines = lines[:max_lines]
        if lines:
            lines[-1] = textwrap.shorten(lines[-1], width=max(8, len(lines[-1]) - 3), placeholder="...")

    total_h = len(lines) * line_height
    if valign == "top":
        first_baseline = y_top - pad - size
    elif valign == "bottom":
        first_baseline = y + pad + total_h - size
    else:
        first_baseline = y + (h + total_h) / 2 - size

    for i, line in enumerate(lines):
        line_y = first_baseline - i * line_height
        if align == "center":
            line_x = x + w / 2
            c.drawCentredString(line_x, line_y, line)
        elif align == "right":
            line_x = x + w - pad
            c.drawRightString(line_x, line_y, line)
        else:
            line_x = x + pad
            c.drawString(line_x, line_y, line)


def _col_positions(x: float) -> tuple[list[float], list[float]]:
    widths_mm = [17, 23, 17, 18, 18, 34]
    widths = [v * mm for v in widths_mm]
    positions = [x]
    for width in widths[:-1]:
        positions.append(positions[-1] + width)
    return positions, widths


def _span(positions: list[float], widths: list[float], start_col: int, end_col: int) -> tuple[float, float]:
    return positions[start_col], sum(widths[start_col : end_col + 1])


def draw_span_cell(
    c: canvas.Canvas,
    positions: list[float],
    widths: list[float],
    start_col: int,
    end_col: int,
    y_top: float,
    h: float,
    text: Any = "",
    **kwargs: Any,
) -> None:
    x, w = _span(positions, widths, start_col, end_col)
    draw_cell(c, x, y_top, w, h, text, **kwargs)


def draw_front_card(
    c: canvas.Canvas,
    x: float,
    y_top: float,
    cage: dict[str, Any],
    settings: dict[str, str],
    include_comments: bool,
) -> None:
    rows_mm = [7.5, 5.7, 5.7, 5.7, 6.5, 7.0, 6.2] + [5.45] * MAX_MICE_PER_CAGE
    rows = [v * mm for v in rows_mm]
    cols, widths = _col_positions(x)
    y = y_top

    protocol_num = safe_str(cage.get("protocol_num")) or settings.get("protocol_num", "")
    approved_date = format_date(cage.get("approved_date")) or format_date(settings.get("approved_date", ""))
    expires_date = format_date(cage.get("expires_date")) or format_date(settings.get("expires_date", ""))
    source = safe_str(cage.get("source")) or settings.get("source", "")
    disposition = safe_str(cage.get("disposition")).upper() or "STOCK"
    litter_line = safe_str(cage.get("litter_mouseline")) if disposition == "MATING" else ""

    # Header row.
    h = rows[0]
    draw_span_cell(c, cols, widths, 0, 1, y, h, f"PI: {settings.get('PI_name', '')}", bold=True, size=8.6, align="center", fill=HEADER_FILL, wrap=False)
    draw_span_cell(c, cols, widths, 2, 3, y, h, f"Protocol: {protocol_num}", bold=True, size=8.6, align="center", fill=HEADER_FILL, wrap=False)
    draw_span_cell(c, cols, widths, 4, 5, y, h, f"Cage #: {cage.get('cage_tag', '')}", bold=True, size=8.6, align="center", fill=HEADER_FILL, wrap=False)
    y -= h

    h = rows[1]
    draw_cell(c, cols[0], y, widths[0], h, "Contact", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, settings.get("contact_name", ""), size=7.2)
    draw_cell(c, cols[3], y, widths[3], h, "Approved", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, approved_date, size=7.2, align="center")
    y -= h

    h = rows[2]
    draw_cell(c, cols[0], y, widths[0], h, "Email", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, settings.get("contact_phone", ""), size=7.0)
    draw_cell(c, cols[3], y, widths[3], h, "Expires", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, expires_date, size=7.2, align="center")
    y -= h

    h = rows[3]
    draw_cell(c, cols[0], y, widths[0], h, "Species", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, settings.get("species", "Mouse"), size=7.2)
    draw_cell(c, cols[3], y, widths[3], h, "Source", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, source, size=7.2, align="center")
    y -= h

    h = rows[4]
    draw_cell(c, cols[0], y, widths[0], h, "Status", bold=True, fill=LABEL_FILL)
    status_fill = colors.black if disposition == "MATING" else _hex("#D9D9D9")
    status_text_color = WHITE if disposition == "MATING" else BLACK
    draw_span_cell(c, cols, widths, 1, 2, y, h, disposition, bold=True, size=7.8, align="center", fill=status_fill, text_color=status_text_color, wrap=False)
    draw_cell(c, cols[3], y, widths[3], h, "Litter Line" if disposition == "MATING" else "", bold=True, fill=LABEL_FILL if disposition == "MATING" else None)
    draw_span_cell(c, cols, widths, 4, 5, y, h, litter_line, size=6.6, wrap=True)
    y -= h

    h = rows[5]
    note_text = compact_note(cage.get("comment", ""), max(0, len(cage.get("mice", [])) - MAX_MICE_PER_CAGE)) if include_comments else ""
    draw_cell(c, cols[0], y, widths[0], h, "Notes", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 5, y, h, note_text, size=6.4, valign="top")
    y -= h

    h = rows[6]
    draw_cell(c, cols[0], y, widths[0], h, "Tag", bold=True, align="center", fill=TABLE_FILL)
    draw_cell(c, cols[1], y, widths[1], h, "DOB", bold=True, align="center", fill=TABLE_FILL)
    draw_cell(c, cols[2], y, widths[2], h, "Sex", bold=True, align="center", fill=TABLE_FILL)
    draw_span_cell(c, cols, widths, 3, 5, y, h, "Genotype", bold=True, align="center", fill=TABLE_FILL)
    y -= h

    visible_mice = cage.get("mice", [])[:MAX_MICE_PER_CAGE]
    genotype_lines = list(cage.get("genotypes", []))
    if len(genotype_lines) < len(cage.get("mice", [])):
        genotype_lines.extend([""] * (len(cage.get("mice", [])) - len(genotype_lines)))

    for i in range(MAX_MICE_PER_CAGE):
        h = rows[7 + i]
        if i < len(visible_mice):
            mouse = visible_mice[i]
            genotype = genotype_lines[i] if i < len(genotype_lines) else ""
            draw_cell(c, cols[0], y, widths[0], h, mouse.get("tag", ""), size=6.7)
            draw_cell(c, cols[1], y, widths[1], h, mouse.get("dob", ""), size=6.7, align="center", wrap=False)
            draw_cell(c, cols[2], y, widths[2], h, mouse.get("sex", ""), size=6.7, align="center", wrap=False)
            draw_span_cell(c, cols, widths, 3, 5, y, h, genotype, size=6.3)
        else:
            draw_cell(c, cols[0], y, widths[0], h, "")
            draw_cell(c, cols[1], y, widths[1], h, "")
            draw_cell(c, cols[2], y, widths[2], h, "")
            draw_span_cell(c, cols, widths, 3, 5, y, h, "")
        y -= h

    # Exact-size outer border, even if row rounding leaves a fractional sliver.
    c.setLineWidth(0.75)
    c.setStrokeColor(BLACK)
    c.rect(x, y_top - CARD_H, CARD_W, CARD_H, fill=0, stroke=1)


def draw_back_card(c: canvas.Canvas, x: float, y_top: float, cage: dict[str, Any]) -> None:
    info = cage.get("mating_info") or {}
    if safe_str(cage.get("disposition")) != "mating" or not info:
        return

    rows_mm = [7.5, 6.2, 6.2, 6.2, 8.8, 30.0, 12.1]
    rows = [v * mm for v in rows_mm]
    cols, widths = _col_positions(x)
    y = y_top

    h = rows[0]
    draw_cell(c, x, y, CARD_W, h, f"Mating Info Back | Cage {cage.get('cage_tag', '')}", bold=True, size=9.0, align="center", fill=HEADER_FILL, wrap=False)
    y -= h

    h = rows[1]
    draw_cell(c, cols[0], y, widths[0], h, "Litter Line", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, info.get("litter_mouseline", "") or cage.get("litter_mouseline", ""), size=6.3)
    draw_cell(c, cols[3], y, widths[3], h, "# Litters", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, info.get("num_litters", ""), size=6.8, align="center")
    y -= h

    h = rows[2]
    draw_cell(c, cols[0], y, widths[0], h, "Mating Cage", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 5, y, h, info.get("mating_cage", ""), size=6.8)
    y -= h

    h = rows[3]
    draw_cell(c, cols[0], y, widths[0], h, "Sire Tag", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, info.get("sire_tag", ""), size=6.8)
    draw_cell(c, cols[3], y, widths[3], h, "Dam Tag", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, info.get("dam_tag", ""), size=6.8)
    y -= h

    h = rows[4]
    draw_cell(c, cols[0], y, widths[0], h, "Sire Genotype", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 1, 2, y, h, info.get("sire_genotype", ""), size=5.9, valign="top")
    draw_cell(c, cols[3], y, widths[3], h, "Dam Genotype", bold=True, fill=LABEL_FILL)
    draw_span_cell(c, cols, widths, 4, 5, y, h, info.get("dam_genotype", ""), size=5.9, valign="top")
    y -= h

    h = rows[5]
    draw_cell(c, cols[0], y, widths[0], h, "Litter History", bold=True, fill=LABEL_FILL, valign="top")
    draw_span_cell(c, cols, widths, 1, 5, y, h, info.get("litter_history", ""), size=5.7, valign="top")
    y -= h

    h = rows[6]
    draw_cell(c, cols[0], y, widths[0], h, "Comment", bold=True, fill=LABEL_FILL, valign="top")
    draw_span_cell(c, cols, widths, 1, 5, y, h, info.get("comment", ""), size=5.8, valign="top")

    c.setLineWidth(0.75)
    c.setStrokeColor(BLACK)
    c.rect(x, y_top - CARD_H, CARD_W, CARD_H, fill=0, stroke=1)



def sort_cages_for_print(cages: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return cages in print order: mating cages first, then stock cages.

    Python's sort is stable, so cages keep their original workbook order within
    each group. This makes the default printout more useful while preserving the
    user's row order among mating cages and among stock cages.
    """
    return sorted(cages, key=lambda cage: 0 if safe_str(cage.get("disposition")).lower() == "mating" else 1)

def build_notecards_pdf_bytes(
    xlsx_source: str | Path | bytes | BinaryIO | pd.DataFrame,
    mating_source: str | Path | bytes | BinaryIO | pd.DataFrame | None = None,
    settings: dict[str, Any] | None = None,
    include_comments: bool = True,
) -> tuple[bytes, dict[str, Any]]:
    if REPORTLAB_IMPORT_ERROR is not None:
        raise RuntimeError(
            "ReportLab is required to generate PDFs. Add `reportlab` to requirements.txt "
            "and redeploy/restart the app. Original import error: "
            f"{REPORTLAB_IMPORT_ERROR}"
        )

    settings_norm = normalize_settings(settings)
    cages, warning_messages = load_cages(xlsx_source, mating_source=mating_source)

    # Normalize fallback dates once for rendering.
    settings_norm["approved_date"] = format_date(settings_norm.get("approved_date", ""))
    settings_norm["expires_date"] = format_date(settings_norm.get("expires_date", ""))

    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=PAGE_SIZE)
    c.setTitle("Mouse cage cards")
    c.setAuthor("Mouse cage card generator")

    cages = sort_cages_for_print(cages)
    front_pages = max(1, math.ceil(len(cages) / CARDS_PER_PAGE)) if cages else 0

    for page_idx in range(front_pages):
        group = cages[page_idx * CARDS_PER_PAGE : page_idx * CARDS_PER_PAGE + CARDS_PER_PAGE]

        # Front page: four exact-size card slots, filled left-to-right, top-to-bottom.
        for slot_idx, cage in enumerate(group):
            slot_x, slot_y = CARD_SLOTS[slot_idx]
            draw_front_card(c, slot_x, slot_y, cage, settings_norm, include_comments)
        c.showPage()

        # Back page: mirror left/right within each row so landscape duplex
        # printing with "flip on long edge" places each back behind the
        # correct front card.
        #
        # Front page order: TL, TR, BL, BR
        # Back page order : TR, TL, BR, BL
        #
        # The PDF preview will show backs horizontally swapped by row. That is
        # intentional for this printer/duplex behavior.
        for slot_idx, cage in enumerate(group):
            back_slot_idx = BACK_SLOT_FOR_FRONT_SLOT[slot_idx]
            slot_x, slot_y = CARD_SLOTS[back_slot_idx]
            draw_back_card(c, slot_x, slot_y, cage)
        c.showPage()

    c.save()
    output.seek(0)

    metadata = {
        "num_cards": len(cages),
        "num_front_pages": front_pages,
        "num_pages": front_pages * 2,
        "cards_per_front_page": CARDS_PER_PAGE,
        "sort_order": "mating_first_then_stock",
        "warnings": warning_messages,
        "include_comments": include_comments,
        "card_width_mm": CARD_WIDTH_MM,
        "card_height_mm": CARD_HEIGHT_MM,
        "output_format": "pdf",
        "duplex_mode": "landscape_long_edge_back_pages_left_right_mirrored",
    }
    return output.getvalue(), metadata


# Backward-compatible name used by older app.py versions; it now returns PDF bytes.
def build_notecards_bytes(
    xlsx_source: str | Path | bytes | BinaryIO | pd.DataFrame,
    settings: dict[str, Any] | None = None,
    include_comments: bool = True,
    mating_source: str | Path | bytes | BinaryIO | pd.DataFrame | None = None,
) -> tuple[bytes, dict[str, Any]]:
    return build_notecards_pdf_bytes(
        xlsx_source=xlsx_source,
        mating_source=mating_source,
        settings=settings,
        include_comments=include_comments,
    )


def build_notecards_pdf_file(
    xlsx_source: str | Path | bytes | BinaryIO | pd.DataFrame,
    output_path: str | Path,
    mating_source: str | Path | bytes | BinaryIO | pd.DataFrame | None = None,
    settings: dict[str, Any] | None = None,
    include_comments: bool = True,
) -> dict[str, Any]:
    content, metadata = build_notecards_pdf_bytes(
        xlsx_source=xlsx_source,
        mating_source=mating_source,
        settings=settings,
        include_comments=include_comments,
    )
    output_file = Path(output_path)
    output_file.write_bytes(content)
    metadata["output_path"] = str(output_file)
    return metadata


def load_settings_yaml(yaml_source: str | Path | bytes | BinaryIO) -> dict[str, str]:
    if hasattr(yaml_source, "read"):
        raw = yaml_source.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
    else:
        raw = Path(yaml_source).read_text(encoding="utf-8") if not isinstance(yaml_source, bytes) else yaml_source.decode("utf-8")
    parsed = yaml.safe_load(raw) or {}
    if not isinstance(parsed, dict):
        raise ValueError("settings.yaml must parse to a key/value mapping")
    return normalize_settings(parsed)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate print-ready mouse cage card PDFs from a SoftMouse workbook.")
    parser.add_argument("--input", default="cages.xlsx", help="Path to the cage workbook (.xlsx)")
    parser.add_argument("--mating-input", default="", help="Optional path to the mating workbook (.xlsx)")
    parser.add_argument("--settings-yaml", default="settings.yaml", help="Path to the YAML settings file")
    parser.add_argument("--output", default="notecards.pdf", help="Path for the generated output PDF")
    parser.add_argument(
        "--exclude-comments",
        action="store_true",
        help="Leave the Notes row blank instead of printing spreadsheet comments",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    settings = load_settings_yaml(args.settings_yaml)
    mating_source = args.mating_input or None
    metadata = build_notecards_pdf_file(
        xlsx_source=args.input,
        mating_source=mating_source,
        output_path=args.output,
        settings=settings,
        include_comments=not args.exclude_comments,
    )

    print("--------------------------------------")
    print(f"Printed {metadata['num_cards']} cage card(s) in PDF order.")
    print(f"Estimated PDF pages: {metadata['num_pages']} ({metadata['num_front_pages']} front page(s) plus matching back page(s)).")
    print(f"Layout: {metadata['cards_per_front_page']} card(s) per front page; mating cages first, then stock cages.")
    print(f"Card size: {metadata['card_width_mm']} mm x {metadata['card_height_mm']} mm")
    if metadata["warnings"]:
        print("Warnings:")
        for item in metadata["warnings"]:
            print(f"- {item}")
    print(f"Saved: {metadata['output_path']}")
    print("Print at Actual Size / 100% scale. Use double-sided, flip on long edge.")
    print("Note: back pages are intentionally rotated in the PDF preview so long-edge duplex prints upright behind the correct card.")
    print("--------------------------------------")


if __name__ == "__main__":
    main()
